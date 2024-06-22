from flask import Flask, request, render_template, send_file
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font
from decimal import Decimal
import os

app = Flask(__name__)

# Load the Excel file
file_path = 'rate_sheet.xlsx'

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/calculate', methods=['POST'])
def calculate():
    try:
        user_code = int(request.form['user_code'])
        username = request.form['username']
        r = float(request.form['fat'])
        col = float(request.form['SNF'])
        amount = float(request.form['amount'])
        
        excel_data = pd.ExcelFile(file_path)
        sheet_name = excel_data.sheet_names[0]
        workbook = load_workbook(file_path)
        sheet = workbook[sheet_name]

        def subtract_decimal_column(b, col):
            c = Decimal(str(col)) - Decimal(str(b))
            adjusted_col = int((c * 10) + 12)
            return adjusted_col

        def subtract_decimal_row(a, r):
            r_dec = Decimal(str(r)) - Decimal(str(a))
            adjusted_row = int((r_dec * 10) + 2)
            return adjusted_row  

        def get_cell_value(sheet, row, column):
            cell_value = sheet.cell(row=row, column=column).value
            return round(cell_value, 2) if cell_value is not None else 0

        def get_price(amount, rate):
            Price = round((amount * rate), 2)
            return Price

        b = 8.0
        a = 3.0

        adjusted_col = subtract_decimal_column(b, col)
        adjusted_row = subtract_decimal_row(a, r)
        rate = get_cell_value(sheet, adjusted_row, adjusted_col)
        price = get_price(amount, rate)

        # Create the Excel bill
        bill_file_path = create_excel_bill(user_code=user_code, username=username, fat=r, SNF=col, amount=amount, rate=rate, price=price)

        return send_file(bill_file_path, as_attachment=True)

    except Exception as e:
        return str(e)

def create_excel_bill(user_code, username, fat, SNF, amount, rate, price):
    wb = Workbook()
    ws = wb.active
    ws.title = "Bill"

    # Set the page layout for 4-inch bill paper
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT

    # Set column widths and row heights to fit the bill format
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20

    # Header
    header_font = Font(size=14, bold=True)
    ws.append(["Milk Purchase Bill"])
    ws["A1"].font = header_font
    ws.merge_cells("A1:B1")
    ws["A1"].alignment = Alignment(horizontal="center")

    # Content
    content_font = Font(size=12)
    ws.append(["User Code", user_code])
    ws.append(["Username", username])
    ws.append(["Fat", fat])
    ws.append(["SNF", SNF])
    ws.append(["Total Amount of Milk", f"{amount} liters"])
    ws.append(["Rate of Milk", f"Rs {rate:.2f} per liter"])
    ws.append(["Total Price", f"Rs {price}"])

    for row in ws.iter_rows(min_row=2, max_row=8, min_col=1, max_col=2):
        for cell in row:
            cell.font = content_font
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # Save the file
    bill_file_path = 'bill.xlsx'
    wb.save(bill_file_path)
    return bill_file_path

if __name__ == '__main__':
    app.run(debug=True)
